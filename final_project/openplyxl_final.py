import os  #os모듈이 필요한 이유 : 정해진 경로의 폴더내에 있는 엑셀파일의 이름을 가져오기 위해 필요
from openpyxl import Workbook #엑셀 파일 생성
from openpyxl import load_workbook #엑셀 파일 열기
import pandas as pd
import numpy as np
from pathlib import Path

# # 해당이름의 csv파일을 읽어옴
r_csv = pd.read_csv("C:/Users/user/Desktop/data/feeds.csv")
# r_csv = r_csv.rename(columns={'field1':'temp'})

# print(r_csv['created_at'])
start_date = str(input("시작하고자 하는 날짜를 입력해주세요. => "))
start_day = start_date.split("-")
start_day1=int(start_day[2])
start_day2=("{:02d}".format(start_day1))
print(start_day2)
start_day3=int(start_day2)
rd_csv= r_csv[r_csv['created_at']> '{}'.format(start_date)].copy()
rd_csv['Date'] = pd.to_datetime(rd_csv['created_at']).dt.normalize()
# print(rd_csv)
# # 저장할 xlsx파일의 이름을 정함
# save_xlsx = pd.read_excel("C:/Users/codka/OneDrive/바탕 화면/data/feedsdhzn.xlsx")
# r_csv.to_excel(save_xlsx, index = False) # xlsx 파일로 변환
# save_xlsx.save() #xlsx 파일로 저장



wb = load_workbook("C:/Users/user/Desktop/data/원예작물재배및실습_11주차_4조_형식자료.xlsx")
ws = wb.active
# print(ws.max_row)
for i in range(2,ws.max_row+1):
    ws.cell(i,1).value = None
    ws.cell(i,2).value = None
    ws.cell(i,3).value = None
    ws.cell(i,4).value = None
    ws.cell(i,5).value = None
    ws.cell(i,7).value = None
    ws.cell(i,8).value = None
    ws.cell(i,9).value = None


line = 0
for i, row in rd_csv.iterrows():
    ws.cell(line+2,1).value = row['created_at']
    ws.cell(line+2,2).value = row['entry_id']
    ws.cell(line+2,3).value = row['field1']
    ws.cell(line+2,4).value = row['field2']
    ws.cell(line+2,5).value = row['field3']
    ws.cell(line+2,7).value = "=LEFT(A{}, 10)".format(line+2)
    ws.cell(line + 2, 8).value = "=0.61078*EXP(C{}/(C{}+233.3)*17.2694)".format(line + 2, line + 2)
    ws.cell(line + 2, 9).value = "=H{}*(1-D{}/100)".format(line + 2, line + 2)

    line +=1
for i in range(7):
    ws.cell(3+i,11).value = "2022-06-{}".format(start_day3+i)
    date = "2022-06-{}".format(start_day3+i)
    # print(date)
    # print(rd_csv[rd_csv['Date'] == '2022-05-29'])
    # print(rd_csv[rd_csv['Date'] == date])
    # print(rd_csv[rd_csv['Date'] == date].max()["field1"])
    ws.cell(3 + i, 12).value = rd_csv[rd_csv['Date'] == date].max()["field1"]
    ws.cell(3 + i, 14).value = rd_csv[rd_csv['Date'] == date].min()["field1"]
    ws.cell(3 + i, 15).value = rd_csv[rd_csv['Date'] == date].max()["field2"]
    ws.cell(3 + i, 17).value = rd_csv[rd_csv['Date'] == date].min()["field2"]
    ws.cell(3 + i, 18).value = rd_csv[rd_csv['Date'] == date].max()["field3"]
    ws.cell(3 + i, 20).value = rd_csv[rd_csv['Date'] == date].min()["field3"]
    # ws.cell(3+i,12).value = "=MAXIFS(C:C,G:G,K{})".format(3+i)
    ws.cell(3+i,13).value                                        = "=AVERAGEIFS(C:C,G:G,K{})".format(3+i)
    # ws.cell(3+i,14).value = "=MINIFS(C:C,G:G,K{})".format(3+i)
    # ws.cell(3+i,15).value = "=MAXIFS(D:D,G:G,K{})".format(3+i)
    ws.cell(3+i,16).value = "=AVERAGEIFS(D:D,G:G,K{})".format(3+i)
    # ws.cell(3+i,17).value = "=MINIFS(D:D,G:G,K{})".format(3+i)
    # ws.cell(3+i,18).value = "=MAXIFS(E:E,G:G,K{})".format(3+i)
    ws.cell(3+i,19).value = "=AVERAGEIFS(E:E,G:G,K{})".format(3+i)
    # ws.cell(3+i,20).value = "=MINIFS  (E:E,G:G,K{})".format(3+i)
    ws.cell(3+i,22).value = "=(S{} * 0.023 * 60 * 60 * 24) / 1000000".format(3+i)  #DLI

for i in range(6):
    ws.cell(3,23).value=""
    ws.cell(4+i,23).value = "=W{}-4.4+M{}".format(3+i, 4+i)    #GDD


week = (int(input("몇주차 자료입니까? =>")))

wb.save('C:/Users/user/Desktop/data/원예작물재배및실습_{}주차_4조_dataclear.xlsx'.format(week))

#df = pd.read_excel('C:/Users/user/Desktop/data/원예작물재배및실습_11주차_4조_dataclear.xlsx')
#print(df)


